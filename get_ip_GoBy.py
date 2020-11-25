"""
get ip , ports ,post agreement from GoBy report.
liushui at HZ

age
python3 get_ip_GoBy.py filename
"""
import sys
import openpyxl


def get_new_sheet(num, str1, str2, str3):  # 将传入的数据添加到新建的sheet中
    #    print(str1.value,str2.value,str3.value)
    cell2 = str2.value
    cell3 = str3.value
    list_2 = cell2.split(',')
    list_3 = cell3.split(',')
    for i in range(0, len(list_2)):
        if (num == 0 or list_2[i] == '-'):  # 判断是否为0或“-”，调整输入格式用，直接将port输出为数字格式
            new_sheet.append({'A': str1.value, 'B': list_2[i], 'C': list_3[i]})
        else:
            new_sheet.append({'A': str1.value, 'B': int(list_2[i]), 'C': list_3[i]})
        print("已写入{}  A:{},B:{},C:{}".format(num, str1.value, str2.value, str3.value))


if __name__ == '__main__':
    name = sys.argv[1]  # 写入表文件名
    #       name = '123456.xlsx'
    web1 = openpyxl.load_workbook(filename=name)  # 加载文件，创建对象
    print("========================================")
    #    print(web1.sheetnames)
    GoBy_sheet = web1['assetSheet']
    new_sheet = web1.create_sheet('port_list')
    A_list = GoBy_sheet['A']
    B_list = GoBy_sheet['B']
    C_list = GoBy_sheet['C']

    for i in range(0, len(A_list)):  # 按列读取文件内容
        get_new_sheet(num=i, str1=A_list[i], str2=B_list[i], str3=C_list[i])
    #        print(A_list[i].value)
    #        print(GoBy_sheet.cell(row = i+1, column = 1).value)

    web1.save(filename=name)
    print("已完成表{}".format(web1.sheetnames[-1]))
    print("========================================")
